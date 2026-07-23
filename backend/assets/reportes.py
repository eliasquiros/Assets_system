"""Reporte de auditoria en XLSX (RF-006).

Genera un libro Excel con una hoja por categoria: cada activo con todos sus
datos (menos id/version/fecha_creacion), y con la depreciacion acumulada, el
valor en libros y el estado recalculados al corte fiscal del 30 de septiembre
del anio solicitado (RN-001). Los activos cuya fecha de inicio de uso es
posterior al corte se excluyen: a esa fecha aun no existian en operacion.
"""
import calendar
import re
from datetime import date

from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import ValidationError
from rest_framework.views import APIView

from .depreciacion import calcular_depreciacion
from .models import Activo

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

# Encabezados de columna, en el orden en que se escriben. La categoria no va
# como columna: es el nombre de la hoja (y el titulo de la fila 1).
COLUMNAS = [
    'N.º de activo', 'Nombre', 'Costo original', 'F. adquisición', 'F. inicio de uso',
    'Vida útil (años)', 'Estado', 'Valor en libros', 'Dep. acumulada',
    'Serie', 'Factura', 'Área', 'Proveedor', 'Marca', 'Modelo', 'Origen',
]
# Caracteres prohibidos por Excel en el nombre de una hoja.
_TITULO_INVALIDO = re.compile(r'[\[\]:*?/\\]')

# Formatos y ancho por columna (indice 1-based, en el orden de COLUMNAS).
# 'fecha' -> DD/MM/YYYY; 'moneda' -> 2 decimales con separador de miles.
_FMT_FECHA = 'DD/MM/YYYY'
_FMT_MONEDA = '#,##0.00'
FORMATOS_COLUMNA = {3: _FMT_MONEDA, 4: _FMT_FECHA, 5: _FMT_FECHA,
                    8: _FMT_MONEDA, 9: _FMT_MONEDA}
# Ancho minimo/maximo (en caracteres) al autoajustar segun el contenido.
_ANCHO_MIN = 10
_ANCHO_MAX = 35
_ENCABEZADO_FILL = PatternFill('solid', fgColor='1F4E78')  # azul corporativo
_ENCABEZADO_FONT = Font(bold=True, color='FFFFFF')
_CENTRADO = Alignment(horizontal='center', vertical='center')


def _nombre_hoja(nombre, usados):
    """Nombre de hoja valido para Excel: sin caracteres prohibidos, <= 31 chars
    y unico dentro del libro (desambigua con un sufijo si ya se uso)."""
    base = _TITULO_INVALIDO.sub(' ', nombre).strip()[:31] or 'Sin categoría'
    candidato = base
    i = 2
    while candidato.lower() in usados:
        sufijo = f' ({i})'
        candidato = base[:31 - len(sufijo)] + sufijo
        i += 1
    usados.add(candidato.lower())
    return candidato


def _validar_anio(valor):
    try:
        anio = int(valor)
    except (TypeError, ValueError):
        raise ValidationError({'anio': 'Debe ser un año válido (YYYY).'})
    if not (2000 <= anio <= 2100):
        raise ValidationError({'anio': 'Fuera de rango (2000–2100).'})
    return anio


def _fila(activo, corte):
    """Los 16 valores de un activo, con dep/libros/estado recalculados al corte."""
    dep, libros, estado = calcular_depreciacion(
        activo.costo_original, activo.vida_util_anios, activo.fecha_inicio, hasta=corte,
    )
    return [
        activo.numero_activo, activo.nombre, float(activo.costo_original),
        activo.fecha_adquisicion, activo.fecha_inicio, activo.vida_util_anios,
        # Etiqueta amigable del estado recalculado al corte (no el almacenado).
        dict(Activo.ESTADO_CHOICES)[estado],
        float(libros), float(dep),
        activo.serie or '', activo.factura or '',
        activo.localizacion.nombre,
        activo.proveedor.nombre if activo.proveedor_id else '',
        activo.marca.nombre if activo.marca_id else '',
        activo.modelo.nombre if activo.modelo_id else '',
        activo.origen.nombre if activo.origen_id else '',
    ]


def construir_libro(anio):
    """Arma el Workbook del reporte de auditoria al 30/09 del `anio`."""
    corte = date(anio, 9, 30)
    activos = (
        Activo.objects
        .filter(fecha_inicio__lte=corte)
        .select_related('localizacion', 'categoria', 'proveedor', 'marca', 'modelo', 'origen')
        .order_by('categoria__nombre', 'numero_activo')
    )

    # Agrupa preservando el orden por nombre de categoria.
    grupos = {}
    for activo in activos:
        grupos.setdefault(activo.categoria.nombre, []).append(activo)

    wb = Workbook()
    wb.remove(wb.active)  # quita la hoja vacia por defecto
    usados = set()

    for categoria, items in grupos.items():
        ws = wb.create_sheet(title=_nombre_hoja(categoria, usados))
        ws['A1'] = categoria
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])  # fila 2 en blanco
        ws.append(COLUMNAS)  # fila 3: encabezados
        for celda in ws[ws.max_row]:
            celda.font = _ENCABEZADO_FONT
            celda.fill = _ENCABEZADO_FILL
            celda.alignment = _CENTRADO
        for activo in items:
            ws.append(_fila(activo, corte))
        _dar_formato(ws)

    if not wb.worksheets:
        wb.create_sheet(title='Sin activos')

    return wb


def _ancho_visible(celda):
    """Longitud aproximada de como se ve la celda ya formateada (no el valor crudo)."""
    valor = celda.value
    if valor is None:
        return 0
    if celda.number_format == _FMT_FECHA:
        return len('DD/MM/YYYY')
    if celda.number_format == _FMT_MONEDA:
        return len(f'{valor:,.2f}')
    return len(str(valor))


def _dar_formato(ws):
    """Formato de fechas/moneda, ancho de columna autoajustado al contenido
    (con tope minimo/maximo) y encabezados congelados."""
    # El bloque de datos empieza en la fila 4 (1=titulo, 2=blanco, 3=headers).
    for fila in ws.iter_rows(min_row=4, max_col=len(COLUMNAS)):
        for celda in fila:
            fmt = FORMATOS_COLUMNA.get(celda.column)
            if fmt:
                celda.number_format = fmt

    for i, encabezado in enumerate(COLUMNAS, start=1):
        letra = get_column_letter(i)
        max_dato = max(
            (_ancho_visible(c) for c in ws[letra][3:]), default=0,
        )
        ancho = max(len(encabezado), max_dato) + 2  # margen
        ws.column_dimensions[letra].width = min(max(ancho, _ANCHO_MIN), _ANCHO_MAX)
    ws.freeze_panes = 'A4'  # mantiene titulo + encabezados visibles al scrollear


class ReporteAuditoriaView(APIView):
    """GET /api/reportes/auditoria/?anio=<YYYY> — descarga el .xlsx de auditoria."""

    def get(self, request):
        anio = _validar_anio(request.query_params.get('anio'))
        wb = construir_libro(anio)
        response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
        response['Content-Disposition'] = (
            f'attachment; filename="reporte_auditoria_{anio}.xlsx"'
        )
        wb.save(response)
        return response


# ---------------------------------------------------------------------------
# Reporte financiero (RF-006): "Reporte de valoracion del inventario de activos
# fijos". Una sola hoja con el mismo formato del Excel de referencia: encabezado
# con la empresa que lo solicita y el mes de corte, un resumen ejecutivo por
# categoria y el detalle de activos agrupado por categoria con subtotales y un
# total general. El corte es el ultimo dia del mes solicitado; la depreciacion
# se recalcula a esa fecha y se excluyen los activos cuya fecha de inicio de uso
# es posterior (a esa fecha aun no estaban en operacion).
# ---------------------------------------------------------------------------

# Nombre del mes en espanol (variante costarricense: "setiembre").
_MESES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
    7: 'julio', 8: 'agosto', 9: 'setiembre', 10: 'octubre', 11: 'noviembre',
    12: 'diciembre',
}

# Columnas del detalle, en el mismo orden que el Excel de referencia. El estado
# va al final (columna I), tal como en la referencia.
COLUMNAS_FIN = [
    '# Activo', 'Descripción', 'Fecha de Compra', 'Fecha Inicio',
    'Vida Útil (años)', 'Valor Original (₡)', 'Depreciación Acumulada (₡)',
    'Valor en Libros (₡)', 'Estado',
]
_ANCHO_TOTAL = len(COLUMNAS_FIN)  # A..I

# Formatos del Excel de referencia: moneda en colones sin decimales (negativos
# entre parentesis, cero como "-"), vida util con sufijo "años" y fechas mm-dd-yy.
_FMT_MONEDA_CRC = r'\₡#,##0;\(\₡#,##0\);\₡0'
_FMT_VIDA = r'0\ "años"'
_FMT_FECHA_REF = 'mm-dd-yy'
# Indices (1-based) de las columnas del detalle por tipo de formato.
_DET_MONEDA = (6, 7, 8)
_DET_FECHA = (3, 4)
_DET_VIDA = 5
# Anchos de columna tomados del Excel de referencia. La columna E se comparte
# entre "Vida Útil (años)" (detalle) y "Depreciación Acumulada (₡)" (resumen);
# el ancho por defecto la dejaba cortada y los montos grandes se veian mal, asi
# que se le da un ancho propio (similar al de las demas columnas de moneda).
_ANCHOS_FIN = {1: 10.1, 2: 58.9, 3: 14.8, 4: 15.6, 5: 18.0, 6: 20.2, 7: 24.9, 8: 20.2, 9: 23.0}

# Paleta exacta del Excel de referencia.
_AZUL_TITULO = '1F3864'    # navy: fondo de empresa/total y texto de encabezados
_AZUL_SECCION = '2E5395'   # bandas "RESUMEN EJECUTIVO" / "DETALLE DE ACTIVOS"
_AZUL_HEADER = 'D9E2F3'    # fila de encabezados de columna
_AZUL_BANNER = '8EAADB'    # banda por categoria
_AZUL_SUBTOTAL = 'DCE6F1'  # fila de subtotal
_GRIS_SUB = '595959'
_GRIS_PIE = '808080'
_BLANCO = 'FFFFFF'

_BORDE_SUP = Border(top=Side(style='thin', color=_AZUL_TITULO))

_F_EMPRESA = Font(name='Calibri', size=18, bold=True, color=_BLANCO)
_F_TITULO = Font(name='Calibri', size=14, bold=True, color=_AZUL_TITULO)
_F_SUBTITULO = Font(name='Calibri', size=10, italic=True, color=_GRIS_SUB)
_F_FECHA = Font(name='Calibri', size=10, color=_GRIS_SUB)
_F_SECCION = Font(name='Calibri', size=11, bold=True, color=_BLANCO)
_F_HEADER = Font(name='Calibri', size=10, bold=True, color=_AZUL_TITULO)
_F_HEADER_RESUMEN = Font(name='Calibri', size=9, bold=True, color=_AZUL_TITULO)
_F_BANNER = Font(name='Calibri', size=10, bold=True, color=_AZUL_TITULO)
_F_SUBTOTAL = Font(name='Calibri', size=10, bold=True, color=_AZUL_TITULO)
_F_TOTAL = Font(name='Calibri', size=11, bold=True, color=_BLANCO)
_F_DATO = Font(name='Calibri', size=9)
_F_DATO_RESUMEN = Font(name='Calibri', size=10)
_F_PIE = Font(name='Calibri', size=8, italic=True, color=_GRIS_PIE)

_FILL_EMPRESA = PatternFill('solid', fgColor=_AZUL_TITULO)
_FILL_SECCION = PatternFill('solid', fgColor=_AZUL_SECCION)
_FILL_HEADER = PatternFill('solid', fgColor=_AZUL_HEADER)
_FILL_BANNER = PatternFill('solid', fgColor=_AZUL_BANNER)
_FILL_SUBTOTAL = PatternFill('solid', fgColor=_AZUL_SUBTOTAL)
_FILL_TOTAL = PatternFill('solid', fgColor=_AZUL_TITULO)

_AL_IZQ = Alignment(horizontal='left', vertical='center')
_AL_CEN = Alignment(horizontal='center', vertical='center')
_AL_DER = Alignment(horizontal='right', vertical='center')
_AL_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _validar_corte(valor):
    """Convierte 'YYYY-MM' en el ultimo dia de ese mes (fecha de corte)."""
    m = re.fullmatch(r'(\d{4})-(\d{2})', valor or '')
    if not m:
        raise ValidationError({'corte': 'Indique el mes de corte (YYYY-MM).'})
    anio, mes = int(m.group(1)), int(m.group(2))
    if not (2000 <= anio <= 2100) or not (1 <= mes <= 12):
        raise ValidationError({'corte': 'Mes de corte fuera de rango.'})
    return date(anio, mes, calendar.monthrange(anio, mes)[1])


def _cat_label(cat):
    """Nombre de la categoria con su prefijo, p. ej. 'Equipo de Cómputo (COMP)'."""
    return f'{cat.nombre} ({cat.prefijo})' if cat.prefijo else cat.nombre


def _agrupar_por_categoria(corte):
    """Activos vigentes al corte, agrupados por categoria con dep/libros/estado
    recalculados a esa fecha y los subtotales de cada categoria."""
    activos = (
        Activo.objects
        .filter(fecha_inicio__lte=corte)
        .select_related('categoria')
        .order_by('categoria__nombre', 'numero_activo')
    )
    grupos = {}
    for a in activos:
        dep, libros, estado = calcular_depreciacion(
            a.costo_original, a.vida_util_anios, a.fecha_inicio, hasta=corte,
        )
        g = grupos.setdefault(a.categoria_id, {
            'cat': a.categoria, 'filas': [],
            'orig': 0.0, 'dep': 0.0, 'libros': 0.0,
        })
        g['filas'].append([
            a.numero_activo, a.nombre, a.fecha_adquisicion, a.fecha_inicio,
            a.vida_util_anios, float(a.costo_original), float(dep), float(libros),
            dict(Activo.ESTADO_CHOICES)[estado],
        ])
        g['orig'] += float(a.costo_original)
        g['dep'] += float(dep)
        g['libros'] += float(libros)
    return grupos


def construir_libro_financiero(corte, empresa_nombre):
    """Arma el Workbook del reporte financiero al `corte` (ultimo dia del mes)
    solicitado por `empresa_nombre`, replicando el formato del Excel de
    referencia (paleta, tipografia, bandas por categoria, subtotales y total)."""
    empresa_nombre = empresa_nombre or 'Empresa'
    grupos = _agrupar_por_categoria(corte)
    tot_orig = sum(g['orig'] for g in grupos.values())
    tot_dep = sum(g['dep'] for g in grupos.values())
    tot_libros = sum(g['libros'] for g in grupos.values())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte de Activos Fijos'
    ws.sheet_view.showGridLines = False  # sin cuadricula, como la referencia

    def banda(fila, texto, hasta_col=_ANCHO_TOTAL, font=None, fill=None,
              align=_AL_CEN, borde=None):
        """Fusiona la fila de la columna 1 a `hasta_col` y le pone el texto."""
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=hasta_col)
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.font = font
        celda.alignment = align
        for col in range(1, hasta_col + 1):
            c = ws.cell(row=fila, column=col)
            if fill:
                c.fill = fill
            if borde:
                c.border = borde
        return celda

    def moneda(fila, col, valor, font=_F_DATO):
        celda = ws.cell(row=fila, column=col, value=valor)
        celda.number_format = _FMT_MONEDA_CRC
        celda.font = font
        celda.alignment = _AL_DER
        return celda

    # --- Encabezado: empresa, titulo, subtitulo y mes de corte. Se respetan
    # las filas en blanco (2-3 y 7) del Excel de referencia. ---
    banda(1, empresa_nombre, font=_F_EMPRESA, fill=_FILL_EMPRESA, align=_AL_IZQ)
    banda(4, 'REPORTE DE VALORACIÓN DEL INVENTARIO DE ACTIVOS FIJOS', font=_F_TITULO)
    banda(5, 'Depreciación según contabilidad fiscal — Desglosado por activo',
          font=_F_SUBTITULO)
    banda(6, f'Fecha del reporte: {_MESES[corte.month]} {corte.year}'
             '        |        Moneda: colones costarricenses (₡)', font=_F_FECHA)
    ws.merge_cells('A2:I2')  # espaciador
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[4].height = 20

    # --- Resumen ejecutivo: una fila por categoria + total general. ---
    r = 8
    banda(r, 'RESUMEN EJECUTIVO', font=_F_SECCION, fill=_FILL_SECCION, align=_AL_IZQ)
    ws.row_dimensions[r].height = 20
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for col, texto, align in (
        (1, 'Tipo de Activo', _AL_IZQ), (3, 'Cantidad', _AL_HEADER),
        (4, 'Valor Original (₡)', _AL_HEADER),
        (5, 'Depreciación Acumulada (₡)', _AL_HEADER),
        (6, 'Valor en Libros (₡)', _AL_HEADER),
    ):
        celda = ws.cell(row=r, column=col, value=texto)
        celda.font = _F_HEADER_RESUMEN
        celda.alignment = align
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = _FILL_HEADER
    ws.row_dimensions[r].height = 28

    for g in grupos.values():
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        etiqueta = ws.cell(row=r, column=1, value=_cat_label(g['cat']))
        etiqueta.font = _F_DATO_RESUMEN
        etiqueta.alignment = _AL_IZQ
        cant = ws.cell(row=r, column=3, value=len(g['filas']))
        cant.font = _F_DATO_RESUMEN
        cant.alignment = _AL_CEN
        moneda(r, 4, g['orig'], font=_F_DATO_RESUMEN)
        moneda(r, 5, g['dep'], font=_F_DATO_RESUMEN)
        moneda(r, 6, g['libros'], font=_F_DATO_RESUMEN)

    # Total del resumen: texto navy en negrita, sin relleno y con borde superior.
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    tot = ws.cell(row=r, column=1, value='TOTAL GENERAL')
    tot.font = _F_HEADER  # navy, negrita, tamaño 10
    tot.alignment = _AL_IZQ
    cant = ws.cell(row=r, column=3, value=sum(len(g['filas']) for g in grupos.values()))
    cant.font = _F_HEADER
    cant.alignment = _AL_CEN
    for col, val in ((4, tot_orig), (5, tot_dep), (6, tot_libros)):
        moneda(r, col, val, font=_F_HEADER)
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = _BORDE_SUP

    # --- Detalle de activos, agrupado por categoria. ---
    r += 2
    banda(r, 'DETALLE DE ACTIVOS', font=_F_SECCION, fill=_FILL_SECCION, align=_AL_IZQ)
    ws.row_dimensions[r].height = 20
    r += 1
    for i, texto in enumerate(COLUMNAS_FIN, start=1):
        celda = ws.cell(row=r, column=i, value=texto)
        celda.font = _F_HEADER
        celda.fill = _FILL_HEADER
        celda.alignment = _AL_IZQ if i == 2 else _AL_CEN  # Descripción a la izq.

    for g in grupos.values():
        r += 1
        banda(r, _cat_label(g['cat']).upper(), font=_F_BANNER, fill=_FILL_BANNER,
              align=_AL_IZQ)
        for datos in g['filas']:
            r += 1
            for col, valor in enumerate(datos, start=1):
                celda = ws.cell(row=r, column=col, value=valor)
                celda.font = _F_DATO
                if col in _DET_FECHA:
                    celda.number_format = _FMT_FECHA_REF
                    celda.alignment = _AL_CEN
                elif col == _DET_VIDA:
                    celda.number_format = _FMT_VIDA
                    celda.alignment = _AL_CEN
                elif col in _DET_MONEDA:
                    celda.number_format = _FMT_MONEDA_CRC
                    celda.alignment = _AL_DER
                elif col == 1:  # # Activo
                    celda.alignment = _AL_CEN
                elif col == _ANCHO_TOTAL:  # Estado
                    celda.alignment = _AL_CEN
                else:  # Descripción
                    celda.alignment = _AL_IZQ
        # Subtotal de la categoria: banda clara con borde superior.
        r += 1
        banda(r, f'SUBTOTAL — {g["cat"].nombre.upper()}', hasta_col=5,
              font=_F_SUBTOTAL, fill=_FILL_SUBTOTAL, align=_AL_IZQ, borde=_BORDE_SUP)
        for col, val in ((6, g['orig']), (7, g['dep']), (8, g['libros'])):
            celda = moneda(r, col, val, font=_F_SUBTOTAL)
            celda.fill = _FILL_SUBTOTAL
            celda.border = _BORDE_SUP
        guion = ws.cell(row=r, column=9, value='—')
        guion.font = _F_SUBTOTAL
        guion.alignment = _AL_CEN
        guion.fill = _FILL_SUBTOTAL
        guion.border = _BORDE_SUP
        r += 1  # fila en blanco entre categorias

    # Total general del detalle: banda navy con texto blanco y borde superior.
    banda(r, 'TOTAL GENERAL', hasta_col=5, font=_F_TOTAL, fill=_FILL_TOTAL,
          align=_AL_IZQ, borde=_BORDE_SUP)
    for col, val in ((6, tot_orig), (7, tot_dep), (8, tot_libros)):
        celda = moneda(r, col, val, font=_F_TOTAL)
        celda.fill = _FILL_TOTAL
        celda.border = _BORDE_SUP
    guion = ws.cell(row=r, column=9)
    guion.fill = _FILL_TOTAL
    guion.border = _BORDE_SUP

    # Nota de fuente al pie.
    r += 2
    banda(r, f'Fuente: Sistema de activos fijos de {empresa_nombre}. '
             f'Reporte generado el {date.today().strftime("%d/%m/%Y")}.',
          font=_F_PIE, align=_AL_IZQ)

    for col, ancho in _ANCHOS_FIN.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho

    return wb


class ReporteFinancieroView(APIView):
    """GET /api/reportes/financiero/?corte=<YYYY-MM> — descarga el .xlsx del
    reporte financiero al ultimo dia del mes solicitado."""

    def get(self, request):
        corte = _validar_corte(request.query_params.get('corte'))
        empresa = getattr(connection.tenant, 'nombre', None)
        wb = construir_libro_financiero(corte, empresa)
        response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
        response['Content-Disposition'] = (
            f'attachment; filename="reporte_financiero_{corte:%Y-%m}.xlsx"'
        )
        wb.save(response)
        return response
